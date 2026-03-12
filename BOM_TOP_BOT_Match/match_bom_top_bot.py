from __future__ import annotations

import argparse
import glob
import json
import math
import os
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
from dataclasses import dataclass, field
from typing import Dict, Iterable, List, Tuple

from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font
from openpyxl.utils.cell import range_boundaries

# 엑셀 COM 연동 (Windows, pywin32 필요). 실패 시 GUI에서 버튼 비활성화.
try:
    import win32com.client
    _EXCEL_COM_AVAILABLE = True
except Exception:
    win32com = None  # type: ignore
    _EXCEL_COM_AVAILABLE = False


def _excel_open_file(abs_path: str, app=None):
    """엑셀 프로그램으로 파일을 열고 (app, wb) 반환. app이 있으면 같은 인스턴스에 추가로 열기."""
    if not _EXCEL_COM_AVAILABLE:
        raise RuntimeError("엑셀 연동을 사용하려면 pywin32가 필요합니다. pip install pywin32")
    abs_path = os.path.abspath(abs_path)
    if not os.path.exists(abs_path):
        raise FileNotFoundError(abs_path)
    # 기존 app이 유효한지 검사 (사용자가 엑셀을 닫으면 COM 참조가 무효화됨)
    if app is not None:
        try:
            _ = app.Workbooks
            app.Visible = True
        except Exception:
            app = None
    if app is None:
        app = win32com.client.Dispatch("Excel.Application")
        app.Visible = True
    wb = app.Workbooks.Open(abs_path)
    app.Visible = True
    wb.Activate()
    return app, wb


def _excel_get_selection(app, wb) -> Tuple[str, str]:
    """활성 시트명과 선택 영역 주소(예: B2:B200) 반환. $ 제거. 복수 영역이면 첫 영역만 사용."""
    if not app or not wb:
        raise RuntimeError("먼저 '엑셀에서 열기'로 파일을 열어 주세요.")
    try:
        wb.Activate()
        sheet_name = str(app.ActiveSheet.Name)
        sel = app.Selection
        # win32com에서는 Address가 인자 없이 프로퍼티로 노출됨. $는 제거.
        addr = str(sel.Address).replace("$", "")
        if "," in addr:
            addr = addr.split(",")[0].strip()
        return sheet_name, addr
    except Exception as e:
        raise RuntimeError(f"엑셀 선택을 가져오지 못했습니다: {e}") from e


Key = Tuple[str, str]  # (coord, material)


@dataclass
class SheetData:
    label: str
    sheet_name: str
    key_to_qty: Dict[Key, float] = field(default_factory=dict)
    key_to_rows: Dict[Key, List[int]] = field(default_factory=dict)
    coord_to_count: Dict[str, int] = field(default_factory=dict)
    coord_to_rows: Dict[str, List[int]] = field(default_factory=dict)
    coord_to_material_counts: Dict[str, Dict[str, int]] = field(default_factory=dict)
    material_to_qty: Dict[str, float] = field(default_factory=dict)  # 자재명별 총수량 (분리된 좌표 개수 합)
    qty_mismatch_rows: List[Tuple[int, float, int]] = field(default_factory=list)  # (행번호, 수량열값, 좌표개수) 수량열≠좌표개수인 행


def norm_text(v: object, *, case_insensitive: bool) -> str:
    if v is None:
        return ""
    s = str(v).strip()
    if case_insensitive:
        s = s.lower()
    return s


def norm_qty(v: object) -> float:
    if v is None:
        return 0.0
    if isinstance(v, (int, float)) and not (isinstance(v, float) and math.isnan(v)):
        return float(v)
    s = str(v).strip()
    if not s:
        return 0.0
    s = s.replace(",", "")
    try:
        return float(s)
    except ValueError:
        return 0.0


def split_coords(coord_text: object) -> List[str]:
    if coord_text is None:
        return []
    s = str(coord_text).strip()
    if not s:
        return []
    parts = [p.strip() for p in s.split(",")]
    return [p for p in parts if p]


def parse_single_col_range(addr: str) -> Tuple[int, int, int, int]:
    """
    Returns (min_col, min_row, max_col, max_row). Validates single-column range.
    """
    min_col, min_row, max_col, max_row = range_boundaries(addr)
    if min_col != max_col:
        raise ValueError(f"Range must be single-column: {addr}")
    if max_row < min_row:
        raise ValueError(f"Invalid range rows: {addr}")
    return min_col, min_row, max_col, max_row


def read_col_range_values(ws, addr: str) -> Tuple[List[object], List[int]]:
    min_col, min_row, _max_col, max_row = parse_single_col_range(addr)
    values: List[object] = []
    rows: List[int] = []
    for r in range(min_row, max_row + 1):
        values.append(ws.cell(row=r, column=min_col).value)
        rows.append(r)
    return values, rows


def bump(d: Dict[str, int], k: str, inc: int = 1) -> None:
    d[k] = d.get(k, 0) + inc


def add_row_list(d: Dict, k, row_num: int) -> None:
    if k not in d:
        d[k] = [row_num]
    else:
        d[k].append(row_num)


def build_sheet_data(
    wb,
    *,
    label: str,
    sheet_name: str,
    material_range: str,
    coord_range: str,
    qty_range: str,
    case_insensitive: bool,
) -> SheetData:
    if sheet_name not in wb.sheetnames:
        raise KeyError(f"Sheet not found: {sheet_name}")
    ws = wb[sheet_name]

    mats, mat_rows = read_col_range_values(ws, material_range)
    coords, _coord_rows = read_col_range_values(ws, coord_range)
    qtys, _qty_rows = read_col_range_values(ws, qty_range)

    if not (len(mats) == len(coords) == len(qtys)):
        raise ValueError(
            f"{label}: ranges must have same number of rows "
            f"(mat={len(mats)}, coord={len(coords)}, qty={len(qtys)})"
        )

    sd = SheetData(label=label, sheet_name=sheet_name)

    for i in range(len(mats)):
        mat = norm_text(mats[i], case_insensitive=case_insensitive)
        coord_text = coords[i]
        coords_list = split_coords(coord_text)
        coord_count = len(coords_list)
        qty_col = norm_qty(qtys[i])
        row_num = mat_rows[i]

        if not mat and not coords_list:
            continue

        # 수량 = 구분자로 분리된 좌표명 개수. 수량열 값과 다르면 검증 목록에 기록
        if abs(qty_col - coord_count) > 1e-9:
            sd.qty_mismatch_rows.append((row_num, qty_col, coord_count))

        for coord_raw in coords_list:
            coord = norm_text(coord_raw, case_insensitive=case_insensitive)
            if coord:
                bump(sd.coord_to_count, coord, 1)
                add_row_list(sd.coord_to_rows, coord, row_num)
                if coord not in sd.coord_to_material_counts:
                    sd.coord_to_material_counts[coord] = {}
                mcounts = sd.coord_to_material_counts[coord]
                mcounts[mat] = mcounts.get(mat, 0) + 1

            if coord and mat:
                key: Key = (coord, mat)
                sd.key_to_qty[key] = sd.key_to_qty.get(key, 0.0) + 1.0
                add_row_list(sd.key_to_rows, key, row_num)

    for (_c, mat), q in sd.key_to_qty.items():
        sd.material_to_qty[mat] = sd.material_to_qty.get(mat, 0.0) + q
    return sd


def merge_sheet_data(top: SheetData, bot: SheetData, *, label: str = "TOP+BOT") -> SheetData:
    """TOP과 BOT 시트 데이터를 하나로 합쳐, BOM과의 단일 매칭에 사용할 보드 데이터를 만든다."""
    merged = SheetData(label=label, sheet_name=f"{top.sheet_name}+{bot.sheet_name}")

    for key in set(top.key_to_qty) | set(bot.key_to_qty):
        q = top.key_to_qty.get(key, 0.0) + bot.key_to_qty.get(key, 0.0)
        if q > 0:
            merged.key_to_qty[key] = q
        merged.key_to_rows[key] = (
            top.key_to_rows.get(key, []) + bot.key_to_rows.get(key, [])
        )

    for coord in set(top.coord_to_count) | set(bot.coord_to_count):
        merged.coord_to_count[coord] = top.coord_to_count.get(coord, 0) + bot.coord_to_count.get(coord, 0)
        merged.coord_to_rows[coord] = top.coord_to_rows.get(coord, []) + bot.coord_to_rows.get(coord, [])
        merged.coord_to_material_counts[coord] = {}
        for mat, cnt in top.coord_to_material_counts.get(coord, {}).items():
            merged.coord_to_material_counts[coord][mat] = merged.coord_to_material_counts[coord].get(mat, 0) + cnt
        for mat, cnt in bot.coord_to_material_counts.get(coord, {}).items():
            merged.coord_to_material_counts[coord][mat] = merged.coord_to_material_counts[coord].get(mat, 0) + cnt

    for mat in set(top.material_to_qty) | set(bot.material_to_qty):
        merged.material_to_qty[mat] = top.material_to_qty.get(mat, 0.0) + bot.material_to_qty.get(mat, 0.0)

    merged.qty_mismatch_rows = list(top.qty_mismatch_rows) + list(bot.qty_mismatch_rows)
    return merged


def compute_status(in_bom: bool, in_board: bool, bom_qty: float, board_qty: float) -> str:
    """상태 한글 반환: 한눈에 알아보기 쉽게."""
    if in_bom and in_board:
        if abs(bom_qty - board_qty) < 1e-9:
            return "일치"
        return "수량 불일치"
    if in_bom and not in_board:
        return "TOP/BOT에 없음"
    if (not in_bom) and in_board:
        return "TOP/BOT에만 있음"
    return "알 수 없음"


def _error_type_coord_mode(status: str, in_bom: bool, in_board: bool, mat: str, bom: SheetData, top_bot: SheetData) -> str:
    """좌표+자재 기준 매칭 시 상태 → 불일치 유형 (좌표_미매칭 / 자재_미매칭 / 수량_불일치)."""
    if status == "수량 불일치":
        return "수량_불일치"
    if status == "TOP/BOT에 없음":  # BOM에만 있음
        return "자재_미매칭" if mat not in top_bot.material_to_qty else "좌표_미매칭"
    if status == "TOP/BOT에만 있음":  # 보드에만 있음
        return "자재_미매칭" if mat not in bom.material_to_qty else "좌표_미매칭"
    return "기타"


def _build_error_rows_material(
    match_rows: List[List[object]],
    bom: SheetData,
    top_bot: SheetData,
    *,
    status_idx: int,
    error_type_quantity: str,
    error_type_material: str,
) -> List[List[object]]:
    """자재 기준 매칭 시 에러 데이터 행 생성. 불일치_유형 포함."""
    rows: List[List[object]] = []
    for r in match_rows:
        if r[status_idx] == "일치":
            continue
        mat = r[0]
        status = r[status_idx]
        err_type = "수량_불일치" if status == "수량 불일치" else error_type_material
        coords_bom = sorted({coord for (coord, m) in bom.key_to_qty if m == mat})
        coords_board = sorted({coord for (coord, m) in top_bot.key_to_qty if m == mat})
        coord_list = sorted(set(coords_bom) | set(coords_board))
        coord_str = ", ".join(coord_list) if coord_list else "-"
        rows.append([err_type, coord_str, mat, r[1], r[4], status])
    return rows


def _build_error_rows_coord(
    match_rows: List[List[object]],
    bom: SheetData,
    top_bot: SheetData,
    *,
    status_idx: int,
) -> List[List[object]]:
    """좌표+자재 기준 매칭 시 에러 데이터 행 생성. 불일치_유형(좌표/자재/수량) 포함."""
    rows: List[List[object]] = []
    for r in match_rows:
        if r[status_idx] == "일치":
            continue
        coord, mat = r[1], r[2]
        status = r[status_idx]
        in_bom = (coord, mat) in bom.key_to_qty
        in_tb = (coord, mat) in top_bot.key_to_qty
        err_type = _error_type_coord_mode(status, in_bom, in_tb, mat, bom, top_bot)
        rows.append([err_type, coord, mat, r[3], r[6], status])
    return rows


def ensure_fresh_sheet(wb, name: str):
    if name in wb.sheetnames:
        ws_old = wb[name]
        wb.remove(ws_old)
    return wb.create_sheet(title=name)


def write_table(ws, headers: List[str], rows: Iterable[Iterable[object]]) -> None:
    bold = Font(bold=True)
    ws.append(headers)
    for c in range(1, len(headers) + 1):
        ws.cell(row=1, column=c).font = bold

    for row in rows:
        ws.append(list(row))

    ws.auto_filter.ref = ws.dimensions
    autofit_columns(ws, max_scan_rows=5000)


def autofit_columns(ws, *, max_scan_rows: int) -> None:
    # openpyxl doesn't have true AutoFit; approximate via max string length.
    dims: Dict[int, int] = {}
    max_row = min(ws.max_row, max_scan_rows)
    for r in range(1, max_row + 1):
        for c in range(1, ws.max_column + 1):
            v = ws.cell(row=r, column=c).value
            if v is None:
                continue
            s = str(v)
            dims[c] = max(dims.get(c, 0), len(s))
    for c, ln in dims.items():
        ws.column_dimensions[ws.cell(row=1, column=c).column_letter].width = min(max(10, ln + 2), 60)


def materials_summary(sd: SheetData, coord: str) -> str:
    m = sd.coord_to_material_counts.get(coord, {})
    parts = [f"{k}({v})" for k, v in m.items()]
    return ", ".join(parts)


def rows_summary(rows: List[int]) -> str:
    return ",".join(str(r) for r in rows)


def _load_workbook(path: str):
    if not os.path.exists(path):
        raise FileNotFoundError(path)
    ext = os.path.splitext(path)[1].lower()
    return load_workbook(path, keep_vba=(ext == ".xlsm"))


def run_match(config: dict) -> str:
    """단일 파일(excel_path) 또는 3개 파일(bom_file, top_file, bot_file) 모드 지원."""
    opts = config.get("options", {})
    inplace = bool(opts.get("inplace", False))
    case_insensitive = bool(opts.get("case_insensitive", False))
    material_only_match = bool(opts.get("material_only_match", True))  # 자재명 기준, TOP+BOT 총수량과 BOM 매칭
    output_path = config.get("output_path") or ""

    bom_cfg = config["bom"]
    top_cfg = config["top"]
    bot_cfg = config["bot"]

    # 3개 파일 모드
    if "bom_file" in config:
        bom_path = config["bom_file"]
        top_path = config["top_file"]
        bot_path = config["bot_file"]
        wb_bom = _load_workbook(bom_path)
        wb_top = _load_workbook(top_path)
        wb_bot = _load_workbook(bot_path)
        bom = build_sheet_data(
            wb_bom, label="BOM", sheet_name=bom_cfg["sheet"],
            material_range=bom_cfg["material_range"],
            coord_range=bom_cfg["coord_range"],
            qty_range=bom_cfg["qty_range"],
            case_insensitive=case_insensitive,
        )
        top = build_sheet_data(
            wb_top, label="TOP", sheet_name=top_cfg["sheet"],
            material_range=top_cfg["material_range"],
            coord_range=top_cfg["coord_range"],
            qty_range=top_cfg["qty_range"],
            case_insensitive=case_insensitive,
        )
        bot = build_sheet_data(
            wb_bot, label="BOT", sheet_name=bot_cfg["sheet"],
            material_range=bot_cfg["material_range"],
            coord_range=bot_cfg["coord_range"],
            qty_range=bot_cfg["qty_range"],
            case_insensitive=case_insensitive,
        )
        wb = Workbook()
        default_sheet = wb.active
        if default_sheet:
            wb.remove(default_sheet)
        base_path = bom_path
    else:
        # 단일 파일 모드
        excel_path = config["excel_path"]
        if not os.path.exists(excel_path):
            raise FileNotFoundError(excel_path)
        wb = _load_workbook(excel_path)
        base_path = excel_path
        bom = build_sheet_data(
            wb, label="BOM", sheet_name=bom_cfg["sheet"],
            material_range=bom_cfg["material_range"],
            coord_range=bom_cfg["coord_range"],
            qty_range=bom_cfg["qty_range"],
            case_insensitive=case_insensitive,
        )
        top = build_sheet_data(
            wb, label="TOP", sheet_name=top_cfg["sheet"],
            material_range=top_cfg["material_range"],
            coord_range=top_cfg["coord_range"],
            qty_range=top_cfg["qty_range"],
            case_insensitive=case_insensitive,
        )
        bot = build_sheet_data(
            wb, label="BOT", sheet_name=bot_cfg["sheet"],
            material_range=bot_cfg["material_range"],
            coord_range=bot_cfg["coord_range"],
            qty_range=bot_cfg["qty_range"],
            case_insensitive=case_insensitive,
        )

    # TOP과 BOT을 하나로 합쳐 BOM과만 매칭 (TOP/BOT 따로 보지 않음)
    top_bot = merge_sheet_data(top, bot)

    if material_only_match:
        # 자재명 기준: BOM 수량 vs (TOP+BOT 합산) 수량 매칭
        all_materials = set(bom.material_to_qty) | set(top_bot.material_to_qty)
        match_rows = []
        ok_count = 0
        for mat in sorted(all_materials):
            q_b = bom.material_to_qty.get(mat, 0.0)
            q_t = top.material_to_qty.get(mat, 0.0)
            q_bo = bot.material_to_qty.get(mat, 0.0)
            q_tb = top_bot.material_to_qty.get(mat, 0.0)
            in_b = mat in bom.material_to_qty
            in_tb = mat in top_bot.material_to_qty
            status = compute_status(in_b, in_tb, q_b, q_tb)
            row = [
                mat, q_b, q_t, q_bo, q_tb,
                "Y" if in_b else "", "Y" if in_tb else "",
                status,
            ]
            match_rows.append(row)
            if status == "일치":
                ok_count += 1
        match_headers = ["자재", "BOM수량", "TOP수량", "BOT수량", "TOP+BOT수량", "BOM여부", "TOP+BOT여부", "상태"]
        un_headers = ["자재", "BOM수량", "TOP수량", "BOT수량", "TOP+BOT수량", "상태"]
        unmatched_rows = [r[:5] + [r[7]] for r in match_rows if r[7] != "일치"]
        # 에러 데이터: 불일치 건별로 관련 좌표명 + 불일치 유형(좌표/자재/수량·좌표개수)
        error_data_rows = _build_error_rows_material(
            match_rows, bom, top_bot, status_idx=7,
            error_type_quantity="수량_불일치",
            error_type_material="자재_미매칭",
        )
        error_headers = ["불일치_유형", "불일치_좌표", "자재", "BOM수량", "TOP+BOT수량", "상태", "비고"]
        for row in error_data_rows:
            row.append("")  # 비고
    else:
        # 좌표+자재 기준: (coord, material) 키별로 BOM vs (TOP+BOT 합산) 매칭
        all_keys = set(bom.key_to_qty) | set(top_bot.key_to_qty)
        match_rows = []
        ok_count = 0
        for coord, mat in sorted(all_keys):
            in_b = (coord, mat) in bom.key_to_qty
            in_tb = (coord, mat) in top_bot.key_to_qty
            q_b = bom.key_to_qty.get((coord, mat), 0.0)
            q_t = top.key_to_qty.get((coord, mat), 0.0)
            q_bo = bot.key_to_qty.get((coord, mat), 0.0)
            q_tb = top_bot.key_to_qty.get((coord, mat), 0.0)
            status = compute_status(in_b, in_tb, q_b, q_tb)
            key_str = f"{coord}|{mat}"
            row = [
                key_str, coord, mat, q_b, q_t, q_bo, q_tb,
                "Y" if in_b else "", "Y" if in_tb else "",
                status,
            ]
            match_rows.append(row)
            if status == "일치":
                ok_count += 1
        match_headers = [
            "키(좌표|자재)", "좌표", "자재",
            "BOM수량", "TOP수량", "BOT수량", "TOP+BOT수량",
            "BOM여부", "TOP+BOT여부", "상태",
        ]
        un_headers = [
            "키(좌표|자재)", "좌표", "자재",
            "BOM수량", "TOP수량", "BOT수량", "TOP+BOT수량", "상태",
        ]
        unmatched_rows = [r[:7] + [r[9]] for r in match_rows if r[9] != "일치"]
        # 에러 데이터: 불일치인 (좌표, 자재)별로 좌표명 + 불일치 유형
        error_data_rows = _build_error_rows_coord(match_rows, bom, top_bot, status_idx=9)
        error_headers = ["불일치_유형", "불일치_좌표", "자재", "BOM수량", "TOP+BOT수량", "상태", "비고"]
        for row in error_data_rows:
            row.append("")  # 비고

    unmatched_count = len(unmatched_rows)

    # 수량열값 ≠ 좌표개수 인 행을 에러_데이터에 추가 (불일치_유형: 수량_좌표개수_불일치)
    qty_verify_rows: List[Tuple[str, str, int, float, int]] = []
    for sd in (bom, top, bot):
        for row_num, qty_col, coord_count in sd.qty_mismatch_rows:
            qty_verify_rows.append((sd.label, sd.sheet_name, row_num, qty_col, coord_count))
    for (label, sheet_name, row_num, qty_col, coord_count) in qty_verify_rows:
        error_data_rows.append([
            "수량_좌표개수_불일치", "-", "-", qty_col, coord_count,
            "수량열≠좌표개수", f"시트={label} 시트명={sheet_name} 행={row_num}",
        ])

    ws_match = ensure_fresh_sheet(wb, "매칭결과")
    write_table(ws_match, headers=match_headers, rows=match_rows)

    ws_un = ensure_fresh_sheet(wb, "불일치")
    write_table(ws_un, headers=un_headers, rows=unmatched_rows)

    ws_error = ensure_fresh_sheet(wb, "에러_데이터")
    write_table(ws_error, headers=error_headers, rows=error_data_rows)

    if qty_verify_rows:
        ws_qty = ensure_fresh_sheet(wb, "수량검증")
        write_table(
            ws_qty,
            headers=["시트구분", "시트명", "행번호", "수량열값", "좌표개수"],
            rows=[list(r) for r in qty_verify_rows],
        )

    dup_rows = []
    dup_count_bom = dup_count_top = dup_count_bot = 0
    for sd in (bom, top, bot):
        for coord, cnt in sorted(sd.coord_to_count.items()):
            if cnt > 1:
                if sd.label == "BOM":
                    dup_count_bom += 1
                elif sd.label == "TOP":
                    dup_count_top += 1
                else:
                    dup_count_bot += 1
                dup_rows.append([
                    sd.label, sd.sheet_name, coord, cnt,
                    materials_summary(sd, coord),
                    rows_summary(sd.coord_to_rows.get(coord, [])),
                ])
    ws_dup = ensure_fresh_sheet(wb, "좌표중복")
    write_table(
        ws_dup,
        headers=["시트구분", "시트명", "좌표", "건수", "자재(건수)", "행"],
        rows=dup_rows,
    )

    ws_summary = ensure_fresh_sheet(wb, "요약")
    write_table(
        ws_summary,
        headers=[
            "구분", "매칭됨(일치)", "불일치(TOP+BOT 통합)",
            "중복좌표_BOM", "중복좌표_TOP", "중복좌표_BOT",
        ],
        rows=[
            [
                "BOM vs TOP+BOT",
                ok_count,
                unmatched_count,
                dup_count_bom,
                dup_count_top,
                dup_count_bot,
            ]
        ],
    )

    if "bom_file" in config:
        out = output_path or (os.path.splitext(bom_path)[0] + "_matched.xlsx")
    elif inplace:
        out = base_path
    else:
        if output_path:
            out = output_path
        else:
            root, ext = os.path.splitext(base_path)
            out = f"{root}_matched{ext}" if ext.lower() in (".xlsx", ".xlsm") else f"{root}_matched.xlsx"

    wb.save(out)
    return out


def _get_sheet_names(file_path: str) -> List[str]:
    """엑셀 파일에서 시트 이름 목록 반환. 실패 시 빈 리스트."""
    if not file_path or not os.path.exists(file_path):
        return []
    try:
        wb = load_workbook(file_path, read_only=True)
        names = wb.sheetnames
        wb.close()
        return names
    except Exception:
        return []


def run_gui() -> None:
    """tkinter 폼: BOM/TOP/BOT 파일·시트·범위 선택 후 매칭 실행."""
    root = tk.Tk()
    root.title("BOM / TOP / BOT 매칭")
    root.geometry("820x700")
    root.resizable(True, True)

    main_frame = ttk.Frame(root, padding=12)
    main_frame.pack(fill=tk.BOTH, expand=True)

    # 엑셀 COM 상태 (앱 1개, 워크북 BOM/TOP/BOT 각 1개)
    excel_state: Dict[str, object] = {"app": None, "bom_wb": None, "top_wb": None, "bot_wb": None}

    def open_in_excel(path_var: tk.StringVar, key: str) -> None:
        path = path_var.get().strip()
        if not path:
            messagebox.showwarning("안내", "먼저 파일을 선택해 주세요.")
            return
        try:
            existing_app = excel_state.get("app")
            # 기존 엑셀 인스턴스 유효성 검사 (닫힌 경우 무효화)
            if existing_app is not None:
                try:
                    _ = existing_app.Workbooks
                    existing_app.Visible = True
                except Exception:
                    existing_app = None
                    excel_state["app"] = None
                    excel_state["bom_wb"] = None
                    excel_state["top_wb"] = None
                    excel_state["bot_wb"] = None
            app, wb = _excel_open_file(path, existing_app)
            excel_state["app"] = app
            excel_state[f"{key}_wb"] = wb
            app.Visible = True
            wb.Activate()
            messagebox.showinfo(
                "안내",
                "엑셀에서 파일이 열렸습니다.\n"
                "범위는 시트/범위 영역의 해당 입력란을 더블클릭한 뒤, 엑셀에서 드래그로 선택하고 확인을 누르세요.",
            )
        except Exception as e:
            messagebox.showerror("엑셀 열기 오류", str(e))

    def get_selection_and_apply(wb_key: str, sheet_var: tk.StringVar, range_var: tk.StringVar) -> None:
        app = excel_state.get("app")
        wb = excel_state.get(f"{wb_key}_wb")
        if not app or not wb:
            messagebox.showwarning("안내", "먼저 해당 파일을 '엑셀에서 열기'로 열어 주세요.")
            return
        try:
            sheet_name, addr = _excel_get_selection(app, wb)
            sheet_var.set(sheet_name)
            range_var.set(addr)
        except Exception as e:
            messagebox.showerror("범위 가져오기 오류", str(e))

    def open_range_picker_dialog(
        parent: tk.Widget,
        wb_key: str,
        range_var: tk.StringVar,
        sheet_var: tk.StringVar,
        range_name: str,
    ) -> None:
        """VBA 유저폼 방식: 범위 입력란 더블클릭 시 안내 라벨이 있는 작은 대화상자 → 엑셀에서 선택 후 확인."""
        app = excel_state.get("app")
        wb = excel_state.get(f"{wb_key}_wb")
        if not app or not wb:
            messagebox.showwarning(
                "안내",
                "먼저 해당 파일을 리스트에서 더블클릭하여 엑셀에서 열어 주세요.",
            )
            return
        dlg = tk.Toplevel(parent)
        dlg.title("범위 선택")
        dlg.geometry("420x130")
        dlg.resizable(False, False)
        msg = f"엑셀에서 {range_name} 범위를 드래그로 선택한 뒤 확인을 누르세요."
        ttk.Label(dlg, text=msg, wraplength=380).pack(pady=20, padx=20, fill=tk.X)
        def on_confirm() -> None:
            try:
                sheet_name, addr = _excel_get_selection(app, wb)
                sheet_var.set(sheet_name)
                range_var.set(addr)
                dlg.destroy()
            except Exception as e:
                messagebox.showerror("범위 가져오기 오류", str(e))
        def on_cancel() -> None:
            dlg.destroy()
        btn_frame = ttk.Frame(dlg)
        btn_frame.pack(pady=8)
        ttk.Button(btn_frame, text="확인", command=on_confirm).pack(side=tk.LEFT, padx=4)
        ttk.Button(btn_frame, text="취소", command=on_cancel).pack(side=tk.LEFT, padx=4)
        dlg.bind("<Return>", lambda e: on_confirm())
        dlg.bind("<Escape>", lambda e: on_cancel())
        try:
            app.Visible = True
            wb.Activate()
        except Exception:
            pass
        dlg.transient(root)
        dlg.grab_set()
        dlg.focus_set()

    # 폴더 선택 → 리스트박스 → 더블클릭으로 BOM/TOP/BOT 파일 지정
    folder_path_var = tk.StringVar()
    bom_path_var = tk.StringVar()
    top_path_var = tk.StringVar()
    bot_path_var = tk.StringVar()

    ttk.Label(main_frame, text="파일 폴더:").grid(row=0, column=0, sticky=tk.W, pady=2)
    ttk.Entry(main_frame, textvariable=folder_path_var, width=38).grid(row=0, column=1, padx=4, pady=2)

    def browse_folder() -> None:
        folder = filedialog.askdirectory(title="BOM/TOP/BOT 파일이 있는 폴더 선택")
        if not folder or not os.path.isdir(folder):
            return
        folder_path_var.set(folder)
        file_listbox.delete(0, tk.END)
        for ext in ("*.xlsx", "*.xlsm", "*.xls"):
            for p in sorted(glob.glob(os.path.join(folder, ext))):
                file_listbox.insert(tk.END, os.path.basename(p))

    ttk.Button(main_frame, text="폴더 선택", command=browse_folder).grid(row=0, column=2, pady=2)

    double_click_target = tk.StringVar(value="bom")

    ttk.Label(main_frame, text="더블클릭 시 지정:").grid(row=1, column=0, sticky=tk.W, pady=2)
    frame_radio = ttk.Frame(main_frame)
    frame_radio.grid(row=1, column=1, columnspan=2, sticky=tk.W, padx=4)
    ttk.Radiobutton(frame_radio, text="BOM", variable=double_click_target, value="bom").pack(side=tk.LEFT, padx=4)
    ttk.Radiobutton(frame_radio, text="TOP", variable=double_click_target, value="top").pack(side=tk.LEFT, padx=4)
    ttk.Radiobutton(frame_radio, text="BOT", variable=double_click_target, value="bot").pack(side=tk.LEFT, padx=4)

    list_frame = ttk.Frame(main_frame)
    list_frame.grid(row=2, column=0, columnspan=3, sticky=tk.NSEW, pady=4)
    file_listbox = tk.Listbox(list_frame, height=6, selectmode=tk.SINGLE, font=("Segoe UI", 9))
    scrollbar = ttk.Scrollbar(list_frame)
    file_listbox.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
    scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
    file_listbox.config(yscrollcommand=scrollbar.set)
    scrollbar.config(command=file_listbox.yview)

    path_vars = {"bom": bom_path_var, "top": top_path_var, "bot": bot_path_var}
    sheet_combos = {"bom": None, "top": None, "bot": None}  # 아래에서 콤보 생성 후 채움

    def on_file_double_click(_event) -> None:
        sel = file_listbox.curselection()
        if not sel:
            return
        folder = folder_path_var.get().strip()
        if not folder or not os.path.isdir(folder):
            messagebox.showwarning("안내", "먼저 폴더를 선택해 주세요.")
            return
        name = file_listbox.get(sel[0])
        full_path = os.path.normpath(os.path.join(folder, name))
        if not os.path.isfile(full_path):
            messagebox.showerror("오류", f"파일을 찾을 수 없습니다: {full_path}")
            return
        target = double_click_target.get()
        path_var = path_vars[target]
        path_var.set(full_path)
        combo = sheet_combos.get(target)
        if combo:
            names = _get_sheet_names(full_path)
            combo["values"] = names
            if names:
                combo.set(names[0])
        if _EXCEL_COM_AVAILABLE:
            try:
                open_in_excel(path_var, target)
            except Exception as e:
                messagebox.showerror("엑셀 열기 오류", str(e))

    file_listbox.bind("<Double-Button-1>", on_file_double_click)

    ttk.Label(main_frame, text="저장 경로 (비우면 BOM과 같은 폴더에 _matched.xlsx):").grid(
        row=3, column=0, sticky=tk.W, pady=2
    )
    output_path_var = tk.StringVar()
    ttk.Entry(main_frame, textvariable=output_path_var, width=38).grid(row=3, column=1, padx=4, pady=2)

    def browse_output():
        path = filedialog.asksaveasfilename(
            title="결과 저장 위치",
            defaultextension=".xlsx",
            filetypes=[("Excel", "*.xlsx"), ("모든 파일", "*.*")],
        )
        if path:
            output_path_var.set(path)

    ttk.Button(main_frame, text="찾아보기", command=browse_output).grid(row=3, column=2, pady=2)

    ttk.Separator(main_frame, orient=tk.HORIZONTAL).grid(row=4, column=0, columnspan=4, sticky=tk.EW, pady=10)
    ttk.Label(main_frame, text="시트 및 범위 (직접 입력 또는 범위 입력란 더블클릭 → 엑셀에서 선택 후 확인):").grid(
        row=5, column=0, columnspan=4, sticky=tk.W
    )

    bom_sheet_var = tk.StringVar()
    top_sheet_var = tk.StringVar()
    bot_sheet_var = tk.StringVar()
    bom_mat_var = tk.StringVar(value="B2:B200")
    bom_coord_var = tk.StringVar(value="C2:C200")
    bom_qty_var = tk.StringVar(value="D2:D200")
    top_mat_var = tk.StringVar(value="B2:B200")
    top_coord_var = tk.StringVar(value="C2:C200")
    top_qty_var = tk.StringVar(value="D2:D200")
    bot_mat_var = tk.StringVar(value="B2:B200")
    bot_coord_var = tk.StringVar(value="C2:C200")
    bot_qty_var = tk.StringVar(value="D2:D200")

    def _grid_section(r, label, combo_widget, sheet_var, mat_var, coord_var, qty_var, wb_key: str) -> None:
        ttk.Label(main_frame, text=label).grid(row=r, column=0, sticky=tk.W, pady=2)
        combo_widget.grid(row=r, column=1, padx=4, pady=2)
        ttk.Label(main_frame, text="자재범위").grid(row=r + 1, column=0, sticky=tk.W, pady=2)
        entry_mat = ttk.Entry(main_frame, textvariable=mat_var, width=20)
        entry_mat.grid(row=r + 1, column=1, padx=4, pady=2)
        if _EXCEL_COM_AVAILABLE:
            entry_mat.bind(
                "<Double-Button-1>",
                lambda e, k=wb_key, sv=sheet_var, rv=mat_var: open_range_picker_dialog(
                    main_frame, k, rv, sv, "자재 범위"
                ),
            )
        ttk.Label(main_frame, text="좌표범위").grid(row=r + 2, column=0, sticky=tk.W, pady=2)
        entry_coord = ttk.Entry(main_frame, textvariable=coord_var, width=20)
        entry_coord.grid(row=r + 2, column=1, padx=4, pady=2)
        if _EXCEL_COM_AVAILABLE:
            entry_coord.bind(
                "<Double-Button-1>",
                lambda e, k=wb_key, sv=sheet_var, rv=coord_var: open_range_picker_dialog(
                    main_frame, k, rv, sv, "좌표 범위"
                ),
            )
        ttk.Label(main_frame, text="수량범위").grid(row=r + 3, column=0, sticky=tk.W, pady=2)
        entry_qty = ttk.Entry(main_frame, textvariable=qty_var, width=20)
        entry_qty.grid(row=r + 3, column=1, padx=4, pady=2)
        if _EXCEL_COM_AVAILABLE:
            entry_qty.bind(
                "<Double-Button-1>",
                lambda e, k=wb_key, sv=sheet_var, rv=qty_var: open_range_picker_dialog(
                    main_frame, k, rv, sv, "수량 범위"
                ),
            )

    bom_sheet_combo = ttk.Combobox(main_frame, textvariable=bom_sheet_var, width=20)
    top_sheet_combo = ttk.Combobox(main_frame, textvariable=top_sheet_var, width=20)
    bot_sheet_combo = ttk.Combobox(main_frame, textvariable=bot_sheet_var, width=20)
    sheet_combos["bom"] = bom_sheet_combo
    sheet_combos["top"] = top_sheet_combo
    sheet_combos["bot"] = bot_sheet_combo

    _grid_section(6, "BOM 시트", bom_sheet_combo, bom_sheet_var, bom_mat_var, bom_coord_var, bom_qty_var, "bom")
    _grid_section(10, "TOP 시트", top_sheet_combo, top_sheet_var, top_mat_var, top_coord_var, top_qty_var, "top")
    _grid_section(14, "BOT 시트", bot_sheet_combo, bot_sheet_var, bot_mat_var, bot_coord_var, bot_qty_var, "bot")

    material_only_match_var = tk.BooleanVar(value=True)
    ttk.Checkbutton(
        main_frame,
        text="자재명 기준 매칭 (BOM 수량 vs TOP+BOT 총수량)",
        variable=material_only_match_var,
    ).grid(row=18, column=0, columnspan=3, sticky=tk.W, pady=4)
    case_insensitive_var = tk.BooleanVar(value=False)
    ttk.Checkbutton(
        main_frame, text="좌표/자재 대소문자 무시", variable=case_insensitive_var
    ).grid(row=19, column=0, columnspan=3, sticky=tk.W, pady=4)

    def run_match_from_gui() -> None:
        bom_path = bom_path_var.get().strip()
        top_path = top_path_var.get().strip()
        bot_path = bot_path_var.get().strip()
        if not bom_path or not top_path or not bot_path:
            messagebox.showerror("입력 오류", "BOM, TOP, BOT 파일을 모두 선택해 주세요.")
            return
        if not bom_sheet_var.get().strip():
            messagebox.showerror("입력 오류", "BOM 시트를 선택해 주세요.")
            return
        if not top_sheet_var.get().strip():
            messagebox.showerror("입력 오류", "TOP 시트를 선택해 주세요.")
            return
        if not bot_sheet_var.get().strip():
            messagebox.showerror("입력 오류", "BOT 시트를 선택해 주세요.")
            return
        cfg = {
            "bom_file": bom_path,
            "top_file": top_path,
            "bot_file": bot_path,
            "output_path": output_path_var.get().strip() or None,
            "bom": {
                "sheet": bom_sheet_var.get().strip(),
                "material_range": bom_mat_var.get().strip(),
                "coord_range": bom_coord_var.get().strip(),
                "qty_range": bom_qty_var.get().strip(),
            },
            "top": {
                "sheet": top_sheet_var.get().strip(),
                "material_range": top_mat_var.get().strip(),
                "coord_range": top_coord_var.get().strip(),
                "qty_range": top_qty_var.get().strip(),
            },
            "bot": {
                "sheet": bot_sheet_var.get().strip(),
                "material_range": bot_mat_var.get().strip(),
                "coord_range": bot_coord_var.get().strip(),
                "qty_range": bot_qty_var.get().strip(),
            },
            "options": {
                "inplace": False,
                "material_only_match": material_only_match_var.get(),
                "case_insensitive": case_insensitive_var.get(),
            },
        }
        try:
            out = run_match(cfg)
            messagebox.showinfo("완료", f"결과 저장:\n{out}")
        except Exception as e:
            err_msg = str(e)
            if "ranges must have same number of rows" in err_msg:
                err_msg = (
                    f"{err_msg}\n\n"
                    "자재·좌표·수량 범위는 행 개수가 같아야 합니다.\n"
                    "해당 시트(BOM/TOP/BOT)에서 세 범위를 같은 행 수로 다시 선택해 주세요."
                )
            messagebox.showerror("오류", err_msg)

    ttk.Button(main_frame, text="매칭 실행", command=run_match_from_gui).grid(
        row=20, column=1, columnspan=2, pady=12
    )

    root.mainloop()


def load_config(path: str) -> dict:
    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--config", required=False, help="Path to config JSON (없으면 GUI 실행)")
    args = ap.parse_args()

    if args.config:
        cfg = load_config(args.config)
        out = run_match(cfg)
        print(out)
    else:
        run_gui()


if __name__ == "__main__":
    main()

