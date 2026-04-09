# -*- coding: utf-8 -*-
"""
폴더 구조 생성 (VBA CreatedFolders.bas 이식).
- 기본 경로 아래에 '폴더명'별 상위 폴더를 만들고, 설정된 하위 폴더 이름마다 하위 디렉터리를 만든다.
- 엑셀 복사·워크북 생성은 하지 않는다.
"""

from __future__ import annotations

import json
import re
import shutil
from datetime import date
from pathlib import Path
from typing import Callable

from config import DEFAULT_OUTPUT_DIR, FOLDER_CREATE_SETTINGS_FILE
from openpyxl import Workbook, load_workbook

_ILLEGAL = re.compile(r'[\\/:*?"<>|]')


def replace_illegal_chars(name: str) -> str:
    """Windows 파일명에 쓸 수 없는 문자 제거 (VBA ReplaceIllegalChars 대응)."""
    if not name:
        return ""
    s = _ILLEGAL.sub("", str(name).strip())
    if s.endswith("."):
        s = s.rstrip(".")
    return s.strip()


def load_folder_create_settings() -> dict:
    """기본 경로(str), 하위 폴더 이름 목록(list[str]), 템플릿 엑셀 경로(str), 표지 시트명(str)."""
    default = {
        "base_path": str(DEFAULT_OUTPUT_DIR),
        "subfolders": [],
        "template_xlsx_path": "",
        "cover_sheet_name": "표지",
    }
    try:
        if not FOLDER_CREATE_SETTINGS_FILE.is_file():
            return default
        with FOLDER_CREATE_SETTINGS_FILE.open("r", encoding="utf-8") as f:
            data = json.load(f)
        base = (data.get("base_path") or "").strip()
        subs = data.get("subfolders") or data.get("subfolder_names") or []
        template = (data.get("template_xlsx_path") or data.get("source_template_path") or "").strip()
        cover_sheet = (data.get("cover_sheet_name") or data.get("coverSheet") or "").strip()
        if isinstance(subs, str):
            subs = [ln.strip() for ln in subs.splitlines() if ln.strip()]
        elif isinstance(subs, list):
            subs = [str(x).strip() for x in subs if str(x).strip()]
        else:
            subs = []
        if base:
            default["base_path"] = base
        default["subfolders"] = subs
        default["template_xlsx_path"] = template
        if cover_sheet:
            default["cover_sheet_name"] = cover_sheet
        return default
    except (OSError, json.JSONDecodeError):
        return default


def save_folder_create_settings(
    base_path: str,
    subfolders: list[str],
    *,
    template_xlsx_path: str = "",
    cover_sheet_name: str = "표지",
) -> None:
    data = {
        "base_path": (base_path or "").strip(),
        "subfolders": [s.strip() for s in subfolders if s.strip()],
        "template_xlsx_path": (template_xlsx_path or "").strip(),
        "cover_sheet_name": (cover_sheet_name or "").strip() or "표지",
    }
    FOLDER_CREATE_SETTINGS_FILE.parent.mkdir(parents=True, exist_ok=True)
    with FOLDER_CREATE_SETTINGS_FILE.open("w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)


def ensure_base_directory(base: Path, log: Callable[[str], None] | None = None) -> bool:
    """기본 경로가 없으면 생성 시도."""
    try:
        base.mkdir(parents=True, exist_ok=True)
        return base.is_dir()
    except OSError as e:
        if log:
            log(f"[폴더] 기본 경로 생성 실패: {base} — {e}")
        return False


def create_folder_structure(
    base_path: Path,
    parent_folder_names: list[str],
    subfolder_names: list[str],
    *,
    log: Callable[[str], None] | None = None,
) -> tuple[int, int, list[str]]:
    """
    각 parent 이름마다 base/parent/ 를 만들고, subfolder_names 각각 base/parent/sub/ 생성.
    반환: (성공한 상위 폴더 수, 건너뜀 수, 오류 메시지 목록)
    """
    errs: list[str] = []
    if not ensure_base_directory(base_path, log):
        return 0, len(parent_folder_names), [f"기본 경로를 사용할 수 없습니다: {base_path}"]

    ok_parents = 0
    skipped = 0
    subs_clean = [replace_illegal_chars(s) for s in subfolder_names]
    subs_clean = [s for s in subs_clean if s]

    seen: set[str] = set()
    for raw in parent_folder_names:
        name = replace_illegal_chars(raw)
        if not name:
            skipped += 1
            if log:
                log(f"[폴더] 빈 이름 건너뜀: {raw!r}")
            continue
        key = name.casefold()
        if key in seen:
            skipped += 1
            continue
        seen.add(key)

        parent = base_path / name
        try:
            parent.mkdir(parents=True, exist_ok=True)
            for sub in subs_clean:
                (parent / sub).mkdir(parents=True, exist_ok=True)
            ok_parents += 1
            if log:
                msg = f"[폴더] 생성: {parent}"
                if subs_clean:
                    msg += f" (+ 하위 {len(subs_clean)}개)"
                log(msg)
        except OSError as e:
            errs.append(f"{name}: {e}")
            if log:
                log(f"[폴더] 오류 {name}: {e}")

    return ok_parents, skipped, errs


def create_folder_structure_grouped(
    base_path: Path,
    group_and_parent_names: list[tuple[str, str]],
    subfolder_names: list[str],
    *,
    log: Callable[[str], None] | None = None,
) -> tuple[int, int, list[str]]:
    """
    base/group/parent/ 를 만들고, subfolder_names 각각 base/group/parent/sub/ 생성.

    - group: 예) "고객사_사업명"
    - parent: 예) "폴더명"
    반환: (성공한 (group,parent) 처리 수, 건너뜀 수, 오류 메시지 목록)
    """
    errs: list[str] = []
    if not ensure_base_directory(base_path, log):
        return 0, len(group_and_parent_names), [f"기본 경로를 사용할 수 없습니다: {base_path}"]

    ok = 0
    skipped = 0
    subs_clean = [replace_illegal_chars(s) for s in subfolder_names]
    subs_clean = [s for s in subs_clean if s]

    seen: set[str] = set()
    for raw_group, raw_parent in group_and_parent_names:
        group = replace_illegal_chars(raw_group)
        parent_name = replace_illegal_chars(raw_parent)

        if not group:
            group = "미지정"
        if not parent_name:
            skipped += 1
            if log:
                log(f"[폴더] 빈 폴더명 건너뜀: group={raw_group!r}, parent={raw_parent!r}")
            continue

        key = (group.casefold() + "/" + parent_name.casefold())
        if key in seen:
            skipped += 1
            continue
        seen.add(key)

        parent = base_path / group / parent_name
        try:
            parent.mkdir(parents=True, exist_ok=True)
            for sub in subs_clean:
                (parent / sub).mkdir(parents=True, exist_ok=True)
            ok += 1
            if log:
                msg = f"[폴더] 생성: {parent}"
                if subs_clean:
                    msg += f" (+ 하위 {len(subs_clean)}개)"
                log(msg)
        except OSError as e:
            errs.append(f"{group}/{parent_name}: {e}")
            if log:
                log(f"[폴더] 오류 {group}/{parent_name}: {e}")

    return ok, skipped, errs


def _select_cover_sheet(wb, preferred_name: str) -> object:
    name = (preferred_name or "").strip()
    if name and name in wb.sheetnames:
        return wb[name]
    # fallback: 첫 시트
    return wb[wb.sheetnames[0]]


def copy_template_and_fill_cover(
    *,
    template_xlsx: Path,
    dest_xlsx: Path,
    cover_sheet_name: str,
    p1_to_p9_values: list[str],
    log: Callable[[str], None] | None = None,
) -> tuple[bool, str | None]:
    """
    템플릿 엑셀을 dest_xlsx로 복사한 뒤, 표지 시트 P1~P9에 값을 채운다.
    - p1_to_p9_values: 길이 9 (P1..P9)
    """
    try:
        if not template_xlsx.is_file():
            return False, f"템플릿 파일이 없습니다: {template_xlsx}"
        if dest_xlsx.exists():
            return False, f"이미 존재하여 건너뜀: {dest_xlsx.name}"
        dest_xlsx.parent.mkdir(parents=True, exist_ok=True)
        shutil.copy2(str(template_xlsx), str(dest_xlsx))

        wb = load_workbook(str(dest_xlsx))
        try:
            ws = _select_cover_sheet(wb, cover_sheet_name)
            vals = (p1_to_p9_values + [""] * 9)[:9]
            for i, v in enumerate(vals, start=1):
                ws[f"P{i}"].value = v
            # 템플릿 표지시트 작성일자(F2): 오늘 날짜(YYYY-MM-DD)
            ws["F2"].value = date.today().isoformat()
            wb.save(str(dest_xlsx))
        finally:
            wb.close()

        if log:
            log(f"[파일] 복사+표지입력: {dest_xlsx}")
        return True, None
    except OSError as e:
        return False, str(e)


def create_blank_workbook(
    *,
    dest_xlsx: Path,
    log: Callable[[str], None] | None = None,
) -> tuple[bool, str | None]:
    """
    VBA의 '새 워크북 생성 후 저장'에 해당.
    """
    try:
        if dest_xlsx.exists():
            return False, f"이미 존재하여 건너뜀: {dest_xlsx.name}"
        dest_xlsx.parent.mkdir(parents=True, exist_ok=True)
        wb = Workbook()
        try:
            wb.save(str(dest_xlsx))
        finally:
            wb.close()
        if log:
            log(f"[파일] 새 엑셀 생성: {dest_xlsx}")
        return True, None
    except OSError as e:
        return False, str(e)
