# -*- coding: utf-8 -*-
"""
공정발주 엑셀 통합 로직.
여러 폴더의 .xlsx에서 '작업 발주' 시트 데이터를 읽어 기준 헤더에 맞춰 통합하고,
폴더명/BOM파일명/발행리스트 컬럼에 행 번호 기반 수식을 넣습니다.
"""

from pathlib import Path
import re
from datetime import datetime
from typing import Callable, Optional

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from config import (
    DEFAULT_SOURCE_FOLDER_PATHS,
    SOURCE_PATHS_FILE,
    DEST_SHEET_NAME,
    SOURCE_SHEET_NAME_CONTAINS,
    IGNORE_SHEET_NAME_CONTAINS,
    SOURCE_HEADER_ROW,
    SOURCE_DATA_START_ROW,
    SOURCE_FIRST_COL,
    STANDARD_HEADERS,
)

import json


# --- 헤더 별칭: 소스 시트에 적힌 이름 → 기준 열 번호(1-based) ---
def _build_header_map() -> dict[str, int]:
    """기준 헤더 순서로 열 번호(1-based) 맵을 만들고, 별칭을 등록한다."""
    # 공백/줄바꿈 제거 버전도 키로 쓸 수 있도록
    def norm(s: str) -> str:
        return re.sub(r"[\s\n\r]+", "", s) if s else ""

    mapping: dict[str, int] = {}
    for idx, name in enumerate(STANDARD_HEADERS, start=1):
        mapping[name.strip()] = idx
        mapping[norm(name)] = idx
        # 줄바꿈 변형
        if "\n" in name:
            mapping[name.replace("\n", "\r\n")] = idx

    # VBA와 동일한 별칭
    for std_name, col in list(mapping.items()):
        if "자재입고" in std_name and "수량" in std_name:
            mapping["자재입고\n수량"] = col
            mapping["자재입고수량"] = col
        if "고객사" in std_name and "납품" in std_name:
            mapping["고객사\n납품"] = col
            mapping["고객사납품"] = col
        if std_name == "발주사양":
            mapping["발주사양(생산기술검토)"] = col

    return mapping


def _norm_header(name: str) -> str:
    return (name or "").strip()


def _get_column_indices() -> tuple[int, int, int]:
    """폴더명, BOM파일명, 발행리스트 열 번호(1-based)."""
    try:
        folder_col = STANDARD_HEADERS.index("폴더명") + 1
        bom_col = STANDARD_HEADERS.index("BOM파일명") + 1
        issue_col = STANDARD_HEADERS.index("발행리스트") + 1
        return folder_col, bom_col, issue_col
    except ValueError:
        return 0, 0, 0


def _clean_date_text(raw: str) -> str:
    """날짜 문자열에서 불필요한 부분 제거 및 오타 보정."""
    if not raw:
        return ""
    text = raw
    # 1) 괄호와 괄호 안 텍스트 제거
    text = re.sub(r"\([^)]*\)", "", text)
    text = text.strip()
    # 2) 2026-26-03-01 같은 패턴 보정: '연도-26-월-일' → '연도-월-일'
    m = re.match(r"^(\d{4})-26-(\d{1,2})-(\d{1,2})$", text)
    if m:
        year, month, day = m.groups()
        text = f"{year}-{int(month)}-{int(day)}"
    return text


def _parse_date_from_sheet_and_book(sheet_name: str, book_name: str) -> str:
    """시트명·파일명에서 날짜 문자열 추출 (예: 2025-3-15) 후 클린업."""
    # 연도: 파일명에서 (예: 공정발주25복사본.xlsx → 2025)
    year_str = book_name.replace(".xlsx", "").replace(".xls", "")
    for prefix in ("공정발주", "복사본"):
        year_str = year_str.replace(prefix, "").strip()
    year_str = "20" + year_str[:2] if len(year_str) >= 2 else "20"
    # 시트명: "작업 발주 3월 15일" 등
    add = sheet_name.replace("작업 발주", "").strip()
    for s in ("(조립)", "조립"):
        add = add.replace(s, "").strip()
    add = add.replace("월 ", "-").replace("일", "").strip()
    raw = f"{year_str}-{add}" if add else year_str
    return _clean_date_text(raw)


def _parse_date_for_compare(text: str) -> datetime:
    """정렬/비교용 날짜 파싱. 실패 시 아주 과거 날짜로 처리."""
    cleaned = _clean_date_text(text)
    if not cleaned:
        return datetime.min
    # 일반적인 'YYYY-M-D' 또는 'YYYY-MM-DD' 처리
    try:
        parts = [int(p) for p in cleaned.split("-") if p]
        if len(parts) >= 3:
            year = parts[0]
            # 2026-26-03-01 타입은 이미 _clean_date_text에서 보정됨
            month = parts[1]
            day = parts[2]
            return datetime(year, month, day)
    except Exception:
        pass
    return datetime.min


# 작업지시번호: 문자(영문·한글) 1~2자 + "-" + 숫자4자리 + "-" + 숫자4자리 (예: AB-1234-5678, 지-0001-0002)
_WORK_ORDER_NO_PATTERN = re.compile(r"^[A-Za-z가-힣]{1,2}-\d{4}-\d{4}$")


def _is_valid_work_order_no(val) -> bool:
    if val is None:
        return False
    s = str(val).strip()
    return bool(_WORK_ORDER_NO_PATTERN.fullmatch(s))


def _apply_autofilter_and_style(ws, log: Callable[[str], None], sheet_label: str) -> None:
    """옵션 A: 자동필터 + 헤더 스타일 + 틀 고정 (진짜 테이블 객체 아님)."""
    try:
        mr = max(ws.max_row or 1, 1)
        mc = max(ws.max_column or 1, 1)
        ref = f"A1:{get_column_letter(mc)}{mr}"
        ws.auto_filter.ref = ref
        ws.freeze_panes = "A2"
        fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        for c in range(1, mc + 1):
            cell = ws.cell(row=1, column=c)
            cell.font = Font(bold=True)
            cell.fill = fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        sample_last = min(mr, 500)
        for c in range(1, mc + 1):
            mlen = 10
            for r in range(1, sample_last + 1):
                v = ws.cell(row=r, column=c).value
                if v is not None:
                    mlen = max(mlen, min(len(str(v)), 55))
            ws.column_dimensions[get_column_letter(c)].width = min(mlen * 1.05 + 2, 45)
    except Exception as e:
        log(f"[경고] 시트 스타일 적용 실패 ({sheet_label}): {e}")


def process_folders(
    output_path: Path,
    log: Callable[[str], None],
    source_paths: Optional[list[str]] = None,
) -> bool:
    """
    source_paths 내 모든 .xlsx를 재귀 탐색해 데이터를 통합하고 output_path에 저장한다.
    경로 접근 실패 시 해당 경로만 로그에 남기고 계속 진행한다.
    log(msg)로 진행/오류 메시지를 출력한다.
    반환: 성공 여부.
    """
    # 설정 파일에 정의된 소스 경로가 있으면 우선 사용
    if source_paths is None:
        loaded_paths: list[str] = []
        try:
            if SOURCE_PATHS_FILE.exists():
                with SOURCE_PATHS_FILE.open("r", encoding="utf-8") as f:
                    data = json.load(f)
                    entries = data.get("folders") or data.get("paths") or []
                    loaded_paths = [str(p) for p in entries]
        except Exception as e:
            log(f"[경고] 소스 경로 설정 파일 로드 실패, 기본 경로 사용: {e}")
        source_paths = loaded_paths or DEFAULT_SOURCE_FOLDER_PATHS
    header_map = _build_header_map()
    folder_col, bom_col, issue_col = _get_column_indices()
    num_std_cols = len(STANDARD_HEADERS)
    try:
        job_col_idx = STANDARD_HEADERS.index("작업지시번호") + 1
    except ValueError:
        job_col_idx = 0

    # 통합 데이터: 리스트 of 리스트 (각 행이 한 리스트)
    rows: list[list] = []
    # 1열은 항상 날짜

    def process_one_file(file_path: Path) -> None:
        """
        단일 워크북을 read_only 모드로 열어 iter_rows(values_only=True)로
        한 번에 행 단위로 처리해 성능을 높인다.
        """
        try:
            wb = load_workbook(file_path, read_only=True, data_only=True)
        except Exception as e:
            log(f"  [경고] 파일 열기 실패: {file_path} - {e}")
            return

        try:
            book_name = file_path.name
            for ws in wb.worksheets:
                name = ws.title
                if SOURCE_SHEET_NAME_CONTAINS not in name:
                    continue
                if IGNORE_SHEET_NAME_CONTAINS in name:
                    continue

                # 헤더 + 데이터 전체를 C열부터 values_only로 순회
                header_row_idx = SOURCE_HEADER_ROW
                data_start_row = SOURCE_DATA_START_ROW

                try:
                    row_iter = ws.iter_rows(
                        min_row=header_row_idx,
                        min_col=SOURCE_FIRST_COL,
                        values_only=True,
                    )
                    try:
                        header_row = next(row_iter)
                    except StopIteration:
                        continue
                    headers = [_norm_header(str(v) if v is not None else "") for v in header_row]
                except Exception as e:
                    log(f"  [경고] 시트 헤더 읽기 실패: {name} @ {file_path} - {e}")
                    continue

                # 대상 열 순서
                target_order = []
                for h in headers:
                    col = header_map.get(h) or header_map.get(re.sub(r"[\s\n\r]+", "", h))
                    target_order.append(col if col else 0)

                # 데이터 행: header_row 다음부터 data_start_row까지는 스킵, 이후부터 사용
                add_date = _parse_date_from_sheet_and_book(name, book_name)
                current_row_index = header_row_idx
                for data_row in row_iter:
                    current_row_index += 1
                    if current_row_index < data_start_row:
                        continue
                    if not any(data_row):
                        # 완전히 빈 행이면 건너뜀
                        continue
                    # 1-based 인덱스 사용: row_data[1] = 1열, row_data[num_std_cols] = 마지막 열
                    row_data = [None] * (num_std_cols + 2)
                    row_data[1] = add_date
                    for j, dest_col in enumerate(target_order):
                        if dest_col <= 0:
                            continue
                        if dest_col == folder_col or dest_col == bom_col or dest_col == issue_col:
                            continue
                        if j >= len(data_row):
                            continue
                        val = data_row[j]
                        row_data[dest_col] = val
                    job_val = (
                        row_data[job_col_idx]
                        if job_col_idx and job_col_idx < len(row_data)
                        else None
                    )
                    if not _is_valid_work_order_no(job_val):
                        continue
                    rows.append(row_data)
        finally:
            try:
                wb.close()
            except Exception:
                pass

    for folder_path in source_paths:
        p = Path(folder_path)
        log(f"[폴더 시작] {folder_path}")
        if not p.exists() or not p.is_dir():
            try:
                p.exists()
            except Exception as e:
                log(f"[경로 접근 실패] {folder_path} - {e}")
            else:
                log(f"[경로 없음] {folder_path}")
            continue
        try:
            files = list(p.rglob("*.xlsx"))
        except PermissionError as e:
            log(f"[경로 접근 실패] {folder_path} - {e}")
            continue
        except OSError as e:
            log(f"[경로 접근 실패] {folder_path} - {e}")
            continue

        processed_files = 0
        rows_before_folder = len(rows)
        for f in files:
            if f.suffix.lower() != ".xlsx":
                continue
            # 엑셀 임시 잠금 파일(~$로 시작) 제외
            if f.name.startswith("~$"):
                continue
            process_one_file(f)
            processed_files += 1
            # 너무 자주 찍지 않고, 대략 20개 단위로 진행 상황 로그
            if processed_files % 20 == 0:
                log(f"  [{folder_path}] {processed_files}개 파일 처리 중...")

        rows_after_folder = len(rows)
        added = rows_after_folder - rows_before_folder
        log(f"[폴더 종료] {folder_path} - 파일 {processed_files}개, 추가된 행 {added}개")

    # 1차 후처리: 작업지시번호 기준 최신 날짜만 남기기
    try:
        try:
            job_col = STANDARD_HEADERS.index("작업지시번호") + 1
        except ValueError:
            job_col = 0
        try:
            date_col = STANDARD_HEADERS.index("날짜") + 1
        except ValueError:
            date_col = 1

        latest_by_job: dict[str, tuple[datetime, list]] = {}

        for row in rows:
            # 작업지시번호가 없거나 열 인덱스를 찾지 못한 경우, 최종 결과에서 제외
            if job_col <= 0 or job_col >= len(row):
                continue
            job = row[job_col]
            if not job:
                continue
            date_text = row[date_col] if date_col < len(row) else ""
            dt = _parse_date_for_compare(str(date_text) if date_text is not None else "")
            key = str(job)
            if key not in latest_by_job:
                latest_by_job[key] = (dt, row)
            else:
                if dt >= latest_by_job[key][0]:
                    latest_by_job[key] = (dt, row)

        dedup_rows: list[list] = [pair[1] for pair in latest_by_job.values()]
        rows = dedup_rows
    except Exception as e:
        log(f"[경고] 작업지시번호 중복 제거 중 오류 발생, 원본 전체 사용: {e}")

    # Excel 최대 행 수: 1,048,576 (헤더 1행 + 데이터 1,048,575행)
    EXCEL_MAX_ROWS = 1_048_576
    EXCEL_MAX_DATA_ROWS = EXCEL_MAX_ROWS - 1
    if len(rows) > EXCEL_MAX_DATA_ROWS:
        log(f"[경고] 데이터가 Excel 최대 행 수를 초과합니다. 처음 {EXCEL_MAX_DATA_ROWS}행만 저장합니다. (총 {len(rows)}행)")
        rows = rows[:EXCEL_MAX_DATA_ROWS]

    # 출력: 새 워크북, 1행 헤더, 2행부터 데이터
    try:
        wb_out = Workbook()
        ws_out = wb_out.active
        ws_out.title = DEST_SHEET_NAME
    except Exception as e:
        log(f"[오류] 출력 워크북 생성 실패: {e}")
        return False

    # 1행: 기준 헤더
    for col, h in enumerate(STANDARD_HEADERS, start=1):
        ws_out.cell(row=1, column=col, value=h)

    # 2행부터 데이터 (row_data는 1-based: row_data[1]=1열)
    num_rows_to_write = len(rows)
    for i, row_data in enumerate(rows):
        out_row = i + 2
        for col in range(1, num_std_cols + 1):
            if col < len(row_data) and row_data[col] is not None:
                ws_out.cell(row=out_row, column=col, value=row_data[col])

    # 날짜/고객사납품 컬럼을 날짜 타입으로 정규화
    try:
        try:
            date_col_idx = STANDARD_HEADERS.index("날짜") + 1
        except ValueError:
            date_col_idx = 1
        try:
            cust_ship_idx = STANDARD_HEADERS.index("고객사\n납품") + 1
        except ValueError:
            # 줄바꿈이 제거된 경우 대비
            try:
                cust_ship_idx = STANDARD_HEADERS.index("고객사납품") + 1
            except ValueError:
                cust_ship_idx = 0

        for r in range(2, num_rows_to_write + 2):
            # 날짜 컬럼
            if date_col_idx:
                cell = ws_out.cell(row=r, column=date_col_idx)
                if cell.value:
                    dt = _parse_date_for_compare(str(cell.value))
                    if dt != datetime.min:
                        cell.value = dt
                        cell.number_format = "yyyy-mm-dd"
            # 고객사납품 컬럼
            if cust_ship_idx:
                cell2 = ws_out.cell(row=r, column=cust_ship_idx)
                if cell2.value and not isinstance(cell2.value, datetime):
                    # 텍스트일 가능성이 있을 때만 파싱 시도
                    dt2 = _parse_date_for_compare(str(cell2.value))
                    if dt2 != datetime.min:
                        cell2.value = dt2
                        cell2.number_format = "yyyy-mm-dd"
    except Exception as e:
        log(f"[경고] 날짜 컬럼 타입 정규화 중 오류: {e}")

    # 수식 컬럼 (행 번호 기반)
    # 폴더명 = 품명 & "(" & 품번 & ")"
    # BOM파일명 = 작업지시번호 & " " & 고객사 & "_" & 품명 & "(" & 품번 & ")"
    # 발행리스트 = 사업명 & "-" & 품명 & "(" & 품번 & ")"
    try:
        품명_col = STANDARD_HEADERS.index("품명") + 1
        품번_col = STANDARD_HEADERS.index("품번") + 1
        작업지시번호_col = STANDARD_HEADERS.index("작업지시번호") + 1
        고객사_col = STANDARD_HEADERS.index("고객사") + 1
        사업명_col = STANDARD_HEADERS.index("사업명") + 1
    except ValueError:
        품명_col = 품번_col = 작업지시번호_col = 고객사_col = 사업명_col = 1

    for r in range(2, num_rows_to_write + 2):
        if folder_col:
            ws_out.cell(
                row=r,
                column=folder_col,
                value=f'=IF(OR(ISBLANK({get_column_letter(품명_col)}{r}),ISBLANK({get_column_letter(품번_col)}{r})),"",{get_column_letter(품명_col)}{r}&"("&{get_column_letter(품번_col)}{r}&")")',
            )
        if bom_col:
            ws_out.cell(
                row=r,
                column=bom_col,
                value=f'=IF(OR(ISBLANK({get_column_letter(작업지시번호_col)}{r}),ISBLANK({get_column_letter(고객사_col)}{r})),"",{get_column_letter(작업지시번호_col)}{r}&" "&{get_column_letter(고객사_col)}{r}&"_"&{get_column_letter(품명_col)}{r}&"("&{get_column_letter(품번_col)}{r}&")")',
            )
        if issue_col:
            ws_out.cell(
                row=r,
                column=issue_col,
                value=f'=IF(OR(ISBLANK({get_column_letter(사업명_col)}{r}),ISBLANK({get_column_letter(품명_col)}{r})),"",{get_column_letter(사업명_col)}{r}&"-"&{get_column_letter(품명_col)}{r}&"("&{get_column_letter(품번_col)}{r}&")")',
            )

    # 별도 시트: 파일 생성용 데이터 (날짜, 작업지시번호, 고객사, 폴더명, BOM파일명, 발행리스트)
    try:
        ws_files = wb_out.create_sheet("파일생성용")
        headers_files = ["날짜", "작업지시번호", "고객사", "폴더명", "BOM파일명", "발행리스트"]
        for c, h in enumerate(headers_files, start=1):
            ws_files.cell(row=1, column=c, value=h)

        # 공정발주내역 시트의 각 행을 기반으로 값 계산
        for i in range(num_rows_to_write):
            src_row = i + 2
            out_row = i + 2
            date_val = ws_out.cell(row=src_row, column=date_col_idx).value
            job_val = ws_out.cell(row=src_row, column=작업지시번호_col).value
            cust_val = ws_out.cell(row=src_row, column=고객사_col).value
            name_val = ws_out.cell(row=src_row, column=품명_col).value
            code_val = ws_out.cell(row=src_row, column=품번_col).value
            proj_val = ws_out.cell(row=src_row, column=사업명_col).value

            ws_files.cell(row=out_row, column=1, value=date_val)
            ws_files.cell(row=out_row, column=2, value=job_val)
            ws_files.cell(row=out_row, column=3, value=cust_val)

            # 폴더명, BOM파일명, 발행리스트는 파이썬에서 완성된 문자열로 작성
            if name_val and code_val:
                folder_val = f"{name_val}({code_val})"
            else:
                folder_val = None
            if job_val and cust_val and name_val and code_val:
                bom_val = f"{job_val} {cust_val}_{name_val}({code_val})"
            else:
                bom_val = None
            if proj_val and name_val and code_val:
                issue_val = f"{proj_val}-{name_val}({code_val})"
            else:
                issue_val = None

            ws_files.cell(row=out_row, column=4, value=folder_val)
            ws_files.cell(row=out_row, column=5, value=bom_val)
            ws_files.cell(row=out_row, column=6, value=issue_val)

        # 공정발주내역 시트에서 폴더명/BOM파일명/발행리스트 컬럼 삭제
        # (열 인덱스가 뒤에서 앞으로 당겨지지 않도록 역순으로 삭제)
        cols_to_delete = []
        if folder_col:
            cols_to_delete.append(folder_col)
        if bom_col:
            cols_to_delete.append(bom_col)
        if issue_col:
            cols_to_delete.append(issue_col)
        for col_idx in sorted(cols_to_delete, reverse=True):
            ws_out.delete_cols(col_idx, 1)
    except Exception as e:
        log(f"[경고] 파일생성용 시트 생성 중 오류: {e}")

    _apply_autofilter_and_style(ws_out, log, DEST_SHEET_NAME)
    if "파일생성용" in wb_out.sheetnames:
        _apply_autofilter_and_style(wb_out["파일생성용"], log, "파일생성용")

    try:
        output_path.parent.mkdir(parents=True, exist_ok=True)
        wb_out.save(output_path)
        log(f"저장 완료: {output_path} (총 {len(rows)}행)")
    except Exception as e:
        log(f"[오류] 저장 실패: {output_path} - {e}")
        return False
    return True
